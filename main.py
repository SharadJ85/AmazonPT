import os,requests,openpyxl,datetime,time
from bs4 import BeautifulSoup
from openpyxl.chart import (LineChart, Reference,)



#------------Function to get Amazon site details
def getAmazonPrice(url,cssSelectorTitle,cssSelectorPrice):
    res=requests.get(url)
    stcode=res.status_code
    t=res.text
    soup=BeautifulSoup(t,'lxml')
    if len(soup.select(cssSelectorPrice)) >0:
        Price = soup.select(cssSelectorPrice)[0].text
    else:
        Price = 'Unavailable'

    if  len(soup.select(cssSelectorTitle)) >0:
        Title = soup.select(cssSelectorTitle)[0].text
    else:
        Title= 'Unavailable'
    return stcode,Title.strip(),Price.strip()




#------------Reading the links from Excel
os.chdir('C:\__PriceTracker')#change current directory
wbk=openpyxl.load_workbook('plinks.xlsx')#open plinks excel workbook
fSheet=wbk[wbk.sheetnames[0]]#select 1st sheet
count=fSheet.max_row#total no of rows used
fcolumn = fSheet['A']#select first column of sheet
links=[]#create an empty list
for x in range(count):#range=count
    links.append(fcolumn[x].value)#adds cell value to the list




#------------Use time module
x = datetime.datetime.now()#-time stamp



#------------Writing the values to a new Excel
if not os.path.isdir("C:\__PriceTracker"):#if path doesn't exist
    os.makedirs("C:\__PriceTracker")#create directory

if not os.path.isfile('ProductsPrice'+str(x.strftime("%b_%Y"))+'.xlsx'):
    wb = openpyxl.Workbook()
    dest_filename = 'ProductsPrice'+str(x.strftime("_%b_%Y"))+'.xlsx'
    wb.save(os.path.join('C:\__PriceTracker', dest_filename))

wb = openpyxl.load_workbook(os.path.join('C:\__PriceTracker', dest_filename))
ws =  wb.active
ws.title = 'Products'
#print(wb.sheetnames)
selectSheet = wb['Products']
selectSheet['A1']='Product-Name'
selectSheet['B1']='Price'
selectSheet['C1']='Amazon-Link'
selectSheet['D1']='Res. Status_Code'

for ct in range(count):
    tle = '#productTitle'
    prc = '#priceblock_ourprice'
    url=links[ct]
    s,t,p=getAmazonPrice(url,tle,prc)
    selectSheet['A' + str(ct + 2)] = t
    selectSheet['B' + str(ct + 2)] = p
    selectSheet['C' + str(ct + 2)] = links[ct]
    selectSheet['D' + str(ct + 2)] = s
    ct+=1




#------------Auto adjusting the Excel cell width for readability
size = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            size[cell.column_letter] = len(str(cell.value))+5
for col, value in size.items():
    ws.column_dimensions[col].width = value


"""
#------------Plot the Graph
chart = LineChart()
chart.title = "Amazon Product Prices"
chart.style = 12
chart.y_axis.title = "Price"
chart.y_axis.crossAx = 500
chart.x_axis = DateAxis(crossAx=100)
chart.x_axis.number_format = 'd-mmm'
chart.x_axis.majorTimeUnit = "days"
chart.x_axis.title = "Date"

c2.add_data(data, titles_from_data=True)
dates = Reference(ws, min_col=1, min_row=2, max_row=7)
c2.set_categories(dates)

ws.add_chart(c2, "A61")


"""

#------------Save final changes to the Excel file
wb.save(os.path.join('C:\__PriceTracker', dest_filename))
